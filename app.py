from flask import Flask, request, render_template, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def compare_sheets(file1, file2, result_file, key_column):
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Trim column names to remove leading/trailing spaces
    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    # Ensure the key column exists
    if key_column not in df1.columns or key_column not in df2.columns:
        raise KeyError(f"Both files must contain the key column '{key_column}'.")

    # Check for missing columns
    common_columns = set(df1.columns).intersection(set(df2.columns))
    missing_columns_file1 = set(df2.columns) - set(df1.columns)
    missing_columns_file2 = set(df1.columns) - set(df2.columns)
    if missing_columns_file1 or missing_columns_file2:
        missing_message = []
        if missing_columns_file1:
            missing_message.append(f"Older File is missing columns: {', '.join(missing_columns_file1)}")
        if missing_columns_file2:
            missing_message.append(f"Newer File is missing columns: {', '.join(missing_columns_file2)}")
        raise KeyError(" | ".join(missing_message))

    # Add row numbers to each dataframe
    df1['Row_Number_file1'] = df1.index + 1
    df2['Row_Number_file2'] = df2.index + 1

    # Perform a full outer join on the key column
    merged_df = pd.merge(df1, df2, on=key_column, how='outer', suffixes=('_old', '_new'))

    # Create a new workbook for the result
    result_wb = openpyxl.Workbook()
    result_ws = result_wb.active
    result_ws.title = "Comparison Result"

    # Add a new sheet for the change log
    change_log_ws = result_wb.create_sheet(title="Change Log")

    # Define fill styles for differences and missing rows
    diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    add_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    delete_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Write the headers for the comparison result sheet
    headers = [key_column, "Status", "Row_Number"] + [col for col in df2.columns if col != key_column]
    result_ws.append(headers)

    # Write the headers for the change log sheet
    change_log_ws.append([key_column, "Column", "Old Value", "New Value", "Change Type"])

    # Compare rows and handle missing rows
    for index, row in merged_df.iterrows():
        key_value = row[key_column]
        row_values = [key_value]
        status = ""
        changes = 0

        # Determine if row was added or removed
        row_num_old = int(row['Row_Number_file1']) if pd.notna(row['Row_Number_file1']) else ''
        row_num_new = int(row['Row_Number_file2']) if pd.notna(row['Row_Number_file2']) else ''
        if row_num_old == '':
            status = 'Row was added'
            change_log_ws.append([key_value, '', '', '', status])
        elif row_num_new == '':
            status = 'Row was deleted'
            change_log_ws.append([key_value, '', '', '', status])
        else:
            # Compare each common column
            for col in df2.columns:
                if col == key_column:
                    continue
                value_old = row[f"{col}_old"] if pd.notna(row[f"{col}_old"]) else ''
                value_new = row[f"{col}_new"] if pd.notna(row[f"{col}_new"]) else ''
                if value_old != value_new:
                    changes += 1
                    row_values.append(f"{value_old} | {value_new}")
                    change_log_ws.append([key_value, col, value_old, value_new, "Value has changed"])
                else:
                    row_values.append(value_new)

            if changes > 0:
                status = f"{changes} Changes"
            else:
                status = "0 Changes"

        row_values.insert(1, status)
        row_values.insert(2, f"{row_num_old} | {row_num_new}")

        result_ws.append(row_values)

        # Highlight differences and missing rows
        if status == 'Row was added':
            for col_num in range(1, len(row_values) + 1):
                result_ws.cell(row=index + 2, column=col_num).fill = add_fill
            change_log_ws.cell(row=change_log_ws.max_row, column=5).fill = add_fill
        elif status == 'Row was deleted':
            for col_num in range(1, len(row_values) + 1):
                result_ws.cell(row=index + 2, column=col_num).fill = delete_fill
            change_log_ws.cell(row=change_log_ws.max_row, column=5).fill = delete_fill
        else:
            for col_num, cell_value in enumerate(row_values[3:], start=4):
                result_cell = result_ws.cell(row=index + 2, column=col_num)
                if " | " in str(cell_value):
                    result_cell.fill = diff_fill
            if changes > 0:
                change_log_ws.cell(row=change_log_ws.max_row, column=5).fill = diff_fill

    # Color the "Value has changed" rows in the change log
    for row in change_log_ws.iter_rows(min_row=2, max_row=change_log_ws.max_row, min_col=5, max_col=5):
        for cell in row:
            if cell.value == "Value has changed":
                cell.fill = diff_fill

    result_wb.save(result_file)
    print(f"Comparison complete. Results saved to {result_file}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/compare', methods=['POST'])
def compare():
    if 'file1' not in request.files or 'file2' not in request.files:
        return "No file part"

    file1 = request.files['file1']
    file2 = request.files['file2']
    key_column = request.form['key_column'].strip()

    if file1.filename == '' or file2.filename == '' or key_column == '':
        return "No selected file or key column"

    if file1 and file2 and allowed_file(file1.filename) and allowed_file(file2.filename):
        filename1 = secure_filename(file1.filename)
        filename2 = secure_filename(file2.filename)
        file1_path = os.path.join(app.config['UPLOAD_FOLDER'], filename1)
        file2_path = os.path.join(app.config['UPLOAD_FOLDER'], filename2)
        result_file_path = os.path.join(app.config['UPLOAD_FOLDER'], "comparison_result.xlsx")
        
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        file1.save(file1_path)
        file2.save(file2_path)

        try:
            compare_sheets(file1_path, file2_path, result_file_path, key_column)
        except KeyError as e:
            return str(e)

        response = send_file(result_file_path, as_attachment=True)

        # Cleanup uploaded files and result file
        try:
            os.remove(file1_path)
            os.remove(file2_path)
            os.remove(result_file_path)
        except OSError as e:
            print(f"Error: {e.strerror}")

        return response

    return "Invalid file type"

if __name__ == '__main__':
    app.run(debug=True)
